import openpyxl
from openpyxl.cell.cell import MergedCell
import json
import argparse
import os
import sys
import shutil
from datetime import datetime

# Configurações
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_FILE = os.path.join(BASE_DIR, "TEMPLATE_ANEXO_I.xlsx")
OUTPUT_DIR = os.path.join(BASE_DIR, "Projetos_Gerados_Excel")

def set_cell_value(ws, cell_ref, value):
    """
    Define valor em célula, tratando corretamente células mescladas.
    """
    try:
        cell = ws[cell_ref]
        if isinstance(cell, MergedCell):
            # Encontrar célula mestre do merge
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    master = ws.cell(merged_range.min_row, merged_range.min_col)
                    master.value = value
                    return
        else:
            cell.value = value
    except Exception as e:
        print(f"Erro ao escrever em {cell_ref}: {e}")

def clear_cell_value(ws, row, col):
    """
    Limpa valor de célula, tratando corretamente células mescladas.
    """
    try:
        cell = ws.cell(row=row, column=col)
        if isinstance(cell, MergedCell):
            # Células mescladas não podem ser limpas individualmente
            # Apenas ignorar
            return
        else:
            cell.value = None
    except Exception as e:
        print(f"Erro ao limpar célula ({row},{col}): {e}")

def generate_excel(data, output_path):
    if not os.path.exists(TEMPLATE_FILE):
        print(f"Erro: Template {TEMPLATE_FILE} não encontrado.")
        return False
        
    print(f"Usando template: {TEMPLATE_FILE}")
    
    # Copiar template para destino primeiro
    try:
        shutil.copy(TEMPLATE_FILE, output_path)
    except Exception as e:
        print(f"Erro ao copiar template: {e}")
        return False
        
    try:
        wb = openpyxl.load_workbook(output_path)
        
        # --- Preencher Aba 0 (Unidades Geradoras) ---
        ws0 = wb['0']
        
        # Limpar dados antigos de módulos (linhas 6 a 14, colunas 1-7)
        # CORRIGIDO: Usar função que trata células mescladas
        print("Limpando dados antigos da Aba 0...")
        for r in range(6, 15):
            for c in range(1, 8):
                clear_cell_value(ws0, r, c)
                
        # Preencher Módulos
        # data['modules'] deve ser uma lista de dicts: {potencia, qtd, fabricante, modelo, area}
        current_row = 6
        if 'modules' in data:
            print(f"Preenchendo {len(data['modules'])} módulo(s)...")
            for idx, mod in enumerate(data['modules']):
                pot = float(mod.get('potencia', 0))
                qtd = int(mod.get('qtd', 0))
                area_un = float(mod.get('area', 0))
                
                # Colunas da Aba 0:
                # 1: Item
                # 2: Potência (W)
                # 3: Quantidade
                # 4: Potência Total (kWp)
                # 5: Área Total (m²)
                # 6: Fabricante
                # 7: Modelo
                
                ws0.cell(row=current_row, column=1).value = idx + 1  # Item
                ws0.cell(row=current_row, column=2).value = pot
                ws0.cell(row=current_row, column=3).value = qtd
                ws0.cell(row=current_row, column=4).value = (pot * qtd) / 1000  # kWp
                # Colunas 5, 6, 7 podem ser mescladas - usar set_cell_value
                set_cell_value(ws0, f"E{current_row}", area_un * qtd)  # Área total
                set_cell_value(ws0, f"F{current_row}", mod.get('fabricante', ''))
                set_cell_value(ws0, f"G{current_row}", mod.get('modelo', ''))
                
                print(f"  Módulo {idx+1}: {mod.get('fabricante')} {mod.get('modelo')} - {qtd}x{pot}W")
                current_row += 1
                
        # --- Preencher Aba 1 (Dados Cadastrais) ---
        ws1 = wb['1']
        
        # Mapeamento Expandido
        mapping = {
            'C10': 'nome_cliente',
            'Q10': 'cpf_cnpj',
            'AD10': 'rg',
            'C13': 'endereco',
            'D15': 'cep',
            'I15': 'cidade',
            'N15': 'uf',
            'F29': 'carga_declarada',
            'H29': 'unidade_potencia',
            'C38': 'resp_tecnico_nome',
            'M38': 'resp_tecnico_titulo',
        }
        
        print("Preenchendo Aba 1 (Dados Cadastrais)...")
        for cell_ref, json_key in mapping.items():
            if json_key in data:
                print(f"  {cell_ref} = {data[json_key]}")
                set_cell_value(ws1, cell_ref, data[json_key])
                
        # Salvar
        wb.save(output_path)
        print(f"✓ Excel salvo em: {output_path}")
        return True

    except Exception as e:
        print(f"✗ Erro ao gerar Excel: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    parser = argparse.ArgumentParser(description='Gerador de Excel Equatorial')
    parser.add_argument('--json', type=str, help='String JSON com dados (Legado)')
    parser.add_argument('--input-file', type=str, help='Caminho para arquivo JSON com dados')
    parser.add_argument('--output', type=str, default=OUTPUT_DIR)
    
    args = parser.parse_args()
    
    final_output_dir = args.output
    if not os.path.exists(final_output_dir):
        os.makedirs(final_output_dir)
        
    data = {}
    
    if args.input_file:
        try:
             with open(args.input_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except Exception as e:
            print(f"Erro ao ler arquivo de input: {e}")
            sys.exit(1)
    elif args.json:
        try:
            data = json.loads(args.json)
        except:
            print("Erro JSON String")
            sys.exit(1)
    else:
        # Dados Demo
        data = {
            "nome_cliente": "CLIENTE TESTE EXCEL",
            "cpf_cnpj": "123.456.789-00",
            "rg": "1234567",
            "endereco": "RUA TESTE, 123",
            "cep": "74000-000",
            "cidade": "GOIÂNIA",
            "uf": "GO",
            "carga_declarada": "10",
            "unidade_potencia": "kW",
            "resp_tecnico_nome": "FULVIO FERREIRA BORGES",
            "resp_tecnico_titulo": "Técnico em Eletrotécnica",
            "modules": [
                {"potencia": 550, "qtd": 10, "area": 2.5, "fabricante": "Jinko Solar", "modelo": "JKM550M-72HL4-V"}
            ]
        }
    
    filename = f"Anexo_I_{data.get('nome_cliente', 'Novo').replace(' ', '_')}.xlsx"
    out_path = os.path.join(final_output_dir, filename)
    
    if generate_excel(data, out_path):
        print(f"OUTPUT_EXCEL_PATH={out_path}")
    else:
        sys.exit(1)

if __name__ == "__main__":
    main()
