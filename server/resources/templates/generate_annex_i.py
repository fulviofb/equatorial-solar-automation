
import openpyxl
from openpyxl.cell.cell import MergedCell
import json
import argparse
import os
import sys
import shutil

# Configurações
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_FILE = os.path.join(BASE_DIR, "TEMPLATE_ANEXO_I.xlsx")
OUTPUT_DIR = os.path.join(BASE_DIR, "Projetos_Gerados_Excel")

def set_cell_value(ws, cell_ref, value):
    try:
        if cell_ref not in ws:
             # It might be out of bounds or handled differently, but try accesing to see
             pass
        
        cell = ws[cell_ref]
        if isinstance(cell, MergedCell):
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    master = ws.cell(merged_range.min_row, merged_range.min_col)
                    master.value = value
                    return
        else:
            cell.value = value
    except Exception as e:
        print(f"Erro ao escrever em {cell_ref}: {e}")

def generate_excel(data, output_path):
    if not os.path.exists(TEMPLATE_FILE):
        print(f"Erro: Template {TEMPLATE_FILE} não encontrado.")
        return False
        
    print(f"Usando template: {TEMPLATE_FILE}")
    
    try:
        shutil.copy(TEMPLATE_FILE, output_path)
    except Exception as e:
        print(f"Erro ao copiar template: {e}")
        return False
        
    try:
        wb = openpyxl.load_workbook(output_path)
        
        # --- Aba 0: Unidades Geradoras ---
        if '0' in wb.sheetnames:
            ws0 = wb['0']

            # Modules (Rows 7-16)
            if 'modules' in data and isinstance(data['modules'], list):
                for idx, mod in enumerate(data['modules']):
                    if idx >= 10: break
                    row = 7 + idx
                    ws0.cell(row=row, column=3).value = idx + 1 # Item
                    ws0.cell(row=row, column=4).value = float(mod.get('potencia', 0))
                    ws0.cell(row=row, column=8).value = int(mod.get('qtd', 0))
                    # Area and Total calculated by excel, skipping overwrite unless needed
                    # ws0.cell(row=row, column=16).value = float(mod.get('area', 0)) * int(mod.get('qtd', 0))
                    
                    # Fabricante/Modelo (Cols T, AA)
                    set_cell_value(ws0, f"T{row}", mod.get('fabricante', '').upper())
                    set_cell_value(ws0, f"AA{row}", mod.get('modelo', '').upper())

            # Inverters (Rows 22-51)
            if 'inverters' in data and isinstance(data['inverters'], list):
                for idx, inv in enumerate(data['inverters']):
                    if idx >= 30: break
                    row = 22 + idx
                    ws0.cell(row=row, column=3).value = idx + 1
                    set_cell_value(ws0, f"D{row}", inv.get('fabricante', '').upper())
                    set_cell_value(ws0, f"H{row}", inv.get('modelo', '').upper())
                    set_cell_value(ws0, f"L{row}", float(inv.get('potenciaNominal', inv.get('potencia', 0))))
                    
                    # Technical Data
                    tensao = inv.get('tensao', 220)
                    set_cell_value(ws0, f"P{row}", tensao)
                    set_cell_value(ws0, f"T{row}", float(inv.get('corrente', 0)))
                    set_cell_value(ws0, f"W{row}", inv.get('fatorPotencia', '>0.99'))
                    set_cell_value(ws0, f"Z{row}", inv.get('rendimento', 97))
                    set_cell_value(ws0, f"AC{row}", inv.get('dht', '<3'))

        # --- Aba 1: Dados Cadastrais ---
        if '1' in wb.sheetnames:
            ws1 = wb['1']
            
            mapping = {
                'C10': 'nome_cliente',
                'R10': 'cpf_cnpj',
                'AC9': 'rg',
                'AC10': 'rg_data_emissao',
                'C13': 'endereco',
                'D15': 'cep',
                'I15': 'cidade',
                'Q15': 'uf',
                'V15': 'email',
                'T13': 'celular',
                'F29': 'carga_declarada',
                'P29': 'disjuntor_entrada',
                'F31': 'tipo_ramal',
                'Z17': 'conta_contrato',
                'C38': 'resp_tecnico_nome',
                'M38': 'resp_tecnico_titulo',
                'Y38': 'resp_tecnico_registro',
                'C41': 'resp_tecnico_email',
                'S41': 'resp_tecnico_celular',
                'C44': 'resp_tecnico_endereco', 
                'P44': 'resp_tecnico_cidade',
                'AB43': 'resp_tecnico_uf',
                'G49': 'fonte_primaria', # Validated in Node?
                'G51': 'tipo_geracao',
                'I53': 'modalidade',
                'AC53': 'potencia_total_kw',
                'AC57': 'potencia_inversores_kw'
            }
            
            for cell_ref, key in mapping.items():
                if key in data and data[key] is not None:
                    val = data[key]
                    if isinstance(val, str):
                        val = val.upper()
                    set_cell_value(ws1, cell_ref, val)

        # Save
        if not os.path.exists(OUTPUT_DIR):
            os.makedirs(OUTPUT_DIR)
            
        wb.save(output_path)
        print(f"OUTPUT_EXCEL_PATH={output_path}")
        return True

    except Exception as e:
        print(f"Erro Python: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--input', required=True, help='Path to JSON data file')
    parser.add_argument('--output', required=True, help='Output path')
    args = parser.parse_args()
    
    with open(args.input, 'r', encoding='utf-8') as f:
        data = json.load(f)
        
    out_path = args.output
    # Ensure dir exists
    out_dir = os.path.dirname(out_path)
    if not os.path.exists(out_dir):
        os.makedirs(out_dir)
    
    success = generate_excel(data, out_path)
    if not success:
        sys.exit(1)

if __name__ == "__main__":
    main()
