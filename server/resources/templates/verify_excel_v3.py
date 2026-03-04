import openpyxl

wb = openpyxl.load_workbook("/tmp/test_excel_v3/Anexo_I_TESTE_AUTOMACAO_CORRIGIDO.xlsx")

print("=== ABA 0 (Módulos) - NOVA VERSÃO ===\n")
ws0 = wb['0']

# Verificar cabeçalhos (linha 6)
print("Linha 6 (Cabeçalhos):")
headers = {}
for col_letter, col_num in [('C', 3), ('D', 4), ('H', 8), ('K', 11), ('P', 16), ('T', 20), ('AA', 27)]:
    cell = ws0.cell(6, col_num)
    headers[col_letter] = cell.value
    print(f"  {col_letter}6: {cell.value}")

# Verificar dados (linha 7)
print("\nLinha 7 (Dados do Módulo):")
data_row = {}
for col_letter, col_num in [('C', 3), ('D', 4), ('H', 8), ('K', 11), ('P', 16), ('T', 20), ('AA', 27)]:
    cell = ws0.cell(7, col_num)
    data_row[col_letter] = cell.value
    print(f"  {col_letter}7: {cell.value}")

# Validação
print("\n=== VALIDAÇÃO ===")
validations = {
    'C7 (Item)': data_row['C'] == 1,
    'D7 (Potência)': data_row['D'] == 550,
    'H7 (Quantidade)': data_row['H'] == 10,
    'K7 (Total kWp)': data_row['K'] == 5.5,
    'P7 (Área)': data_row['P'] == 25.0,
    'T7 (Fabricante)': data_row['T'] == 'JINKO SOLAR',
    'AA7 (Modelo)': data_row['AA'] == 'JKM550M-72HL4-V'
}

for field, is_ok in validations.items():
    status = "✓ OK" if is_ok else "✗ FALHA"
    print(f"{field}: {status}")

# Contar sucessos
success_count = sum(validations.values())
total_count = len(validations)
print(f"\nRESULTADO: {success_count}/{total_count} campos corretos ({success_count/total_count*100:.1f}%)")
