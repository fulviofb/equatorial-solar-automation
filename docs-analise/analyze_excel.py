#!/usr/bin/env python3
"""
Script para análise detalhada dos documentos Excel gerados vs templates oficiais da Equatorial GO
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border
import json

def analyze_workbook(filepath, label):
    """Analisa estrutura detalhada de uma planilha Excel"""
    wb = openpyxl.load_workbook(filepath)
    
    analysis = {
        "filepath": filepath,
        "label": label,
        "sheets": []
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        sheet_info = {
            "name": sheet_name,
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "merged_cells": [str(mc) for mc in ws.merged_cells.ranges],
            "sample_cells": []
        }
        
        # Analisar primeiras 50 linhas
        for row_idx in range(1, min(51, ws.max_row + 1)):
            for col_idx in range(1, min(31, ws.max_column + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                
                if cell.value is not None:
                    cell_info = {
                        "address": cell.coordinate,
                        "value": str(cell.value)[:100],  # Limitar tamanho
                        "data_type": cell.data_type,
                    }
                    
                    # Informações de estilo
                    if cell.font:
                        cell_info["font"] = {
                            "bold": cell.font.bold,
                            "size": cell.font.size,
                            "color": str(cell.font.color.rgb) if cell.font.color and hasattr(cell.font.color, 'rgb') else None
                        }
                    
                    if cell.fill and cell.fill.patternType:
                        cell_info["fill"] = {
                            "pattern": cell.fill.patternType,
                            "fgColor": str(cell.fill.fgColor.rgb) if cell.fill.fgColor and hasattr(cell.fill.fgColor, 'rgb') else None
                        }
                    
                    if cell.alignment:
                        cell_info["alignment"] = {
                            "horizontal": cell.alignment.horizontal,
                            "vertical": cell.alignment.vertical,
                            "wrap_text": cell.alignment.wrap_text
                        }
                    
                    sheet_info["sample_cells"].append(cell_info)
        
        analysis["sheets"].append(sheet_info)
    
    return analysis

def compare_structures(template_analysis, generated_analysis):
    """Compara estruturas e identifica diferenças"""
    differences = []
    
    # Comparar número de abas
    template_sheets = {s["name"] for s in template_analysis["sheets"]}
    generated_sheets = {s["name"] for s in generated_analysis["sheets"]}
    
    if template_sheets != generated_sheets:
        differences.append({
            "type": "ESTRUTURA - Abas diferentes",
            "severity": "CRÍTICO",
            "template": list(template_sheets),
            "generated": list(generated_sheets),
            "missing_in_generated": list(template_sheets - generated_sheets),
            "extra_in_generated": list(generated_sheets - template_sheets)
        })
    
    # Comparar cada aba comum
    for template_sheet in template_analysis["sheets"]:
        sheet_name = template_sheet["name"]
        generated_sheet = next((s for s in generated_analysis["sheets"] if s["name"] == sheet_name), None)
        
        if not generated_sheet:
            continue
        
        # Comparar dimensões
        if template_sheet["max_row"] != generated_sheet["max_row"] or \
           template_sheet["max_column"] != generated_sheet["max_column"]:
            differences.append({
                "type": f"ESTRUTURA - Dimensões da aba '{sheet_name}'",
                "severity": "ALTO",
                "template": f"{template_sheet['max_row']}x{template_sheet['max_column']}",
                "generated": f"{generated_sheet['max_row']}x{generated_sheet['max_column']}"
            })
        
        # Comparar células mescladas
        if set(template_sheet["merged_cells"]) != set(generated_sheet["merged_cells"]):
            differences.append({
                "type": f"FORMATAÇÃO - Células mescladas na aba '{sheet_name}'",
                "severity": "MÉDIO",
                "template_count": len(template_sheet["merged_cells"]),
                "generated_count": len(generated_sheet["merged_cells"]),
                "template_sample": template_sheet["merged_cells"][:5],
                "generated_sample": generated_sheet["merged_cells"][:5]
            })
    
    return differences

def main():
    print("=" * 80)
    print("ANÁLISE COMPARATIVA: Documentos Gerados vs Templates Oficiais Equatorial GO")
    print("=" * 80)
    print()
    
    # Analisar template oficial
    print("📋 Analisando template oficial...")
    template_path = "/home/ubuntu/equatorial-solar-automation/server/templates/NT.00020-05-Anexo-I-Formulario-de-Solicitacao-de-Orcamento-de-Microgeracao-Distribuida-Grupo-BHerikson.xlsx"
    template_analysis = analyze_workbook(template_path, "TEMPLATE OFICIAL")
    
    print(f"   ✓ Template possui {len(template_analysis['sheets'])} abas:")
    for sheet in template_analysis["sheets"]:
        print(f"     - {sheet['name']}: {sheet['max_row']} linhas x {sheet['max_column']} colunas")
        print(f"       Células mescladas: {len(sheet['merged_cells'])}")
    print()
    
    # Analisar documento gerado
    print("📄 Analisando documento gerado...")
    generated_path = "/home/ubuntu/equatorial-solar-automation/docs-analise/Formulario_30001.xlsx"
    generated_analysis = analyze_workbook(generated_path, "DOCUMENTO GERADO")
    
    print(f"   ✓ Documento gerado possui {len(generated_analysis['sheets'])} abas:")
    for sheet in generated_analysis["sheets"]:
        print(f"     - {sheet['name']}: {sheet['max_row']} linhas x {sheet['max_column']} colunas")
        print(f"       Células mescladas: {len(sheet['merged_cells'])}")
    print()
    
    # Comparar estruturas
    print("🔍 Identificando diferenças...")
    differences = compare_structures(template_analysis, generated_analysis)
    
    print(f"\n{'=' * 80}")
    print(f"TOTAL DE DIFERENÇAS ENCONTRADAS: {len(differences)}")
    print(f"{'=' * 80}\n")
    
    # Agrupar por severidade
    critical = [d for d in differences if d.get("severity") == "CRÍTICO"]
    high = [d for d in differences if d.get("severity") == "ALTO"]
    medium = [d for d in differences if d.get("severity") == "MÉDIO"]
    
    if critical:
        print(f"🚨 CRÍTICO ({len(critical)}):")
        for diff in critical:
            print(f"\n   {diff['type']}")
            for key, value in diff.items():
                if key not in ["type", "severity"]:
                    print(f"      {key}: {value}")
    
    if high:
        print(f"\n⚠️  ALTO ({len(high)}):")
        for diff in high:
            print(f"\n   {diff['type']}")
            for key, value in diff.items():
                if key not in ["type", "severity"]:
                    print(f"      {key}: {value}")
    
    if medium:
        print(f"\n⚡ MÉDIO ({len(medium)}):")
        for diff in medium:
            print(f"\n   {diff['type']}")
            for key, value in diff.items():
                if key not in ["type", "severity"]:
                    print(f"      {key}: {value}")
    
    # Salvar análise completa
    output = {
        "template": template_analysis,
        "generated": generated_analysis,
        "differences": differences
    }
    
    with open("/home/ubuntu/equatorial-solar-automation/docs-analise/excel_analysis.json", "w", encoding="utf-8") as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    
    print(f"\n{'=' * 80}")
    print("✅ Análise completa salva em: docs-analise/excel_analysis.json")
    print(f"{'=' * 80}\n")

if __name__ == "__main__":
    main()
