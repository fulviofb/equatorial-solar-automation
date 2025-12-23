import openpyxl
from openpyxl.cell.cell import MergedCell
import json

template_path = "../server/python_modules/TEMPLATE_ANEXO_I.xlsx"
generated_path = "Excel_Gerado_V2.xlsx"

print("=" * 80)
print("ANÁLISE COMPARATIVA: Template vs Gerado (V2)")
print("=" * 80)

# Carregar ambos
wb_template = openpyxl.load_workbook(template_path)
wb_generated = openpyxl.load_workbook(generated_path)

# Analisar Aba 0 (Módulos)
print("\n### ABA 0 - UNIDADES GERADORAS ###\n")

ws0_template = wb_template['0']
ws0_generated = wb_generated['0']

print("TEMPLATE - Linha 6 (Exemplo do Adão):")
for c in range(1, 8):
    cell = ws0_template.cell(6, c)
    print(f"  Col {c}: {cell.value}")

print("\nGERADO - Linha 6 (Nosso teste):")
for c in range(1, 8):
    cell = ws0_generated.cell(6, c)
    print(f"  Col {c}: {cell.value}")

# Analisar Aba 1 (Dados Cadastrais)
print("\n### ABA 1 - DADOS CADASTRAIS ###\n")

ws1_template = wb_template['1']
ws1_generated = wb_generated['1']

campos_criticos = {
    'C10': 'Nome Cliente',
    'R10': 'CPF/CNPJ',
    'AC9': 'RG',
    'AC10': 'Data Expedição RG',
    'C13': 'Endereço',
    'D15': 'CEP',
    'I15': 'Cidade',
    'N15': 'UF',
    'F29': 'Carga Declarada',
    'H29': 'Unidade Potência',
    'C38': 'Resp. Técnico Nome',
    'M38': 'Resp. Técnico Título'
}

print("Comparação de Campos:")
print(f"{'Campo':<20} {'Template':<30} {'Gerado':<30} {'Status':<10}")
print("-" * 90)

for cell_ref, descricao in campos_criticos.items():
    template_val = ws1_template[cell_ref].value
    generated_val = ws1_generated[cell_ref].value
    
    # Simplificar valores para comparação
    template_str = str(template_val)[:28] if template_val else "VAZIO"
    generated_str = str(generated_val)[:28] if generated_val else "VAZIO"
    
    status = "✓ OK" if generated_val and generated_val != template_val else "✗ FALHA"
    
    print(f"{descricao:<20} {template_str:<30} {generated_str:<30} {status:<10}")

# Verificar células mescladas problemáticas
print("\n### CÉLULAS MESCLADAS NA ABA 1 ###\n")
print("Verificando se células críticas são mescladas...")

for cell_ref in ['R10', 'AC9', 'AC10']:
    cell = ws1_template[cell_ref]
    is_merged = isinstance(cell, MergedCell)
    if is_merged:
        for merged_range in ws1_template.merged_cells.ranges:
            if cell.coordinate in merged_range:
                master = ws1_template.cell(merged_range.min_row, merged_range.min_col)
                print(f"{cell_ref}: MERGED - Mestre em {master.coordinate} = {master.value}")
                break
    else:
        print(f"{cell_ref}: NORMAL - Valor = {cell.value}")

print("\n" + "=" * 80)
