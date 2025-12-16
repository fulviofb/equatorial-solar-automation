import openpyxl

TEMPLATE_PATH = "c:/Users/fulvi/OneDrive/Agentes - Antigravity/Fóton/temp_analysis/server/python_modules/TEMPLATE_ANEXO_I.xlsx"

def map_cells():
    try:
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        for sheet_name in wb.sheetnames:
            print(f"\n--- Aba: {sheet_name} ---")
            ws = wb[sheet_name]
            # Iterar sobre um range razoável para achar os campos
            for row in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=35):
                for cell in row:
                    if cell.value:
                        val = str(cell.value).strip()
                        if len(val) > 0:
                            print(f"[{cell.coordinate}] = {val[:50]}...")
                            
    except Exception as e:
        print(f"Erro: {e}")

if __name__ == "__main__":
    map_cells()
