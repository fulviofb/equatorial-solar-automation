import openpyxl
import argparse
import json
import sys
import os

def verify_excel(file_path, expectations):
    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        return False
        
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        errors = []
        
        for sheet_name, cells in expectations.items():
            if sheet_name not in wb.sheetnames:
                errors.append(f"Missing sheet: {sheet_name}")
                continue
            
            ws = wb[sheet_name]
            for cell_ref, expected_val in cells.items():
                actual_val = ws[cell_ref].value
                
                # Normalize for comparison
                if isinstance(expected_val, (int, float)) and isinstance(actual_val, (int, float)):
                    # Allow small float diff
                    if abs(expected_val - actual_val) > 0.01:
                         errors.append(f"[{sheet_name}][{cell_ref}] Expected {expected_val}, got {actual_val}")
                else:
                    # String comparison (upper case)
                    exp_str = str(expected_val).strip().upper() if expected_val is not None else ""
                    act_str = str(actual_val).strip().upper() if actual_val is not None else ""
                    
                    if exp_str != act_str:
                        # try numeric comparison if strings look like numbers
                        try:
                            if abs(float(exp_str) - float(act_str)) > 0.01:
                                errors.append(f"[{sheet_name}][{cell_ref}] Expected '{expected_val}', got '{actual_val}'")
                        except:
                            errors.append(f"[{sheet_name}][{cell_ref}] Expected '{expected_val}', got '{actual_val}'")

        if errors:
            print("Verification Failed:")
            for e in errors:
                print(e)
            return False
            
        print("Verification Success")
        return True

    except Exception as e:
        print(f"Error reading excel: {e}")
        return False

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument('--file', required=True)
    parser.add_argument('--expect', required=True)
    args = parser.parse_args()
    
    expect_arg = args.expect
    if os.path.exists(expect_arg):
        with open(expect_arg, 'r', encoding='utf-8') as f:
            expectations = json.load(f)
    else:
        expectations = json.loads(expect_arg)
        
    success = verify_excel(args.file, expectations)
    if not success:
        sys.exit(1)
