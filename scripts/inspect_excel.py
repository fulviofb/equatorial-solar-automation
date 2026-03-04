import openpyxl
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(BASE_DIR, "../server/resources/templates/TEMPLATE_ANEXO_I.xlsx")
OUTPUT_FILE = os.path.join(BASE_DIR, "excel_cells.txt")

def inspect_cells():
    if not os.path.exists(TEMPLATE_PATH):
        print(f"Error: Template not found at {TEMPLATE_PATH}")
        return

    print(f"Inspecting: {TEMPLATE_PATH} -> {OUTPUT_FILE}")
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        for sheet_name in wb.sheetnames:
            f.write(f"\n--- SHEET: {sheet_name} ---\n")
            ws = wb[sheet_name]
            
            # Iterate over a large range to catch everything
            # Sheet 1 has important data
            max_r = 100
            if sheet_name == '1':
                max_r = 150
            
            for row in ws.iter_rows(min_row=1, max_row=max_r, min_col=1, max_col=50):
                for cell in row:
                    if cell.value:
                        val = str(cell.value).strip()
                        if val and len(val) > 0:
                            f.write(f"{cell.coordinate} | {val}\n")

if __name__ == "__main__":
    inspect_cells()
